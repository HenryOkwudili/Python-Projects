#installed pypi for excel implementations
# using pip install openpyxl
import openpyxl as xl #Gives openpyxl an alias xl
from openpyxl.chart import BarChart, Reference

def process_workbook():
    filename = input("Enter the filename: ")
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value*0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

process_workbook()
# process_workbook(filename='pypi.xlsx')